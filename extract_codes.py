import shutil

import openpyxl
import os
from glob import glob

import pandas as pd
from tqdm import tqdm
from datetime import datetime
import ngram

if not os.path.exists('FacturasPendientes'):
    os.mkdir('FacturasPendientes')

if not os.path.exists('FacturasConCodigo'):
    os.mkdir('FacturasConCodigo')

if not os.path.exists('Facturas'):
    os.mkdir('Facturas')


def iter_row_by_column(ws, c):
    datos = []
    for row in ws.iter_rows():
        for col in row:
            if col.coordinate[0].lower() == c.lower():
                if col.value:
                    datos.append(col.value)
    return datos


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row if cell.value]


def make_books():
    doc = openpyxl.load_workbook('CLAVES DEL SAT 2.xlsx')
    for name in doc.sheetnames:
        a = []
        if name not in ['Filtros', 'Envios', 'Factura global']:
            for e in iter_rows(doc[name]):
                if len(e) > 2:
                    e = [e[0], str('').join(e[1:])]
                if len(e) == 2:
                    a.append(e)
                elif len(e) == 1:
                    try:
                        int(e)
                    except Exception:
                        a.append([0, e[0]])
                    else:
                        a.append(e + [""])

            c = pd.DataFrame(a, columns=['SKU', 'Producto'])
            yield c


def get_element_father(z):
    return z.iloc[0, 0], z.iloc[0, 1]


def compare_elements(name_1, n_grama: ngram.NGram):
    res = n_grama.search(name_1)
    for k, v in res:
        if int(v) == 1:
            return True
    return False


def get_by_split(name_product, product_in_sheet, old=None):
    split_name = str(',').join(name_product.split(',')[:-1])
    try:
        split_sheet = product_in_sheet.split(',')
        n_grama = ngram.NGram(split_sheet)
        if len(name_product.split(',')) == 2:
            split_name_2 = str(' ').join(old.split()[:3])
            r = compare_elements(split_name_2, n_grama)
            if r:
                return r
            else:
                get_by_split(split_name, product_in_sheet, old)
        elif len(name_product.split(',')) == 1:
            return False
        r = compare_elements(split_name, n_grama)
        if r:
            return r
        else:
            return get_by_split(split_name, product_in_sheet, name_product)
    except AttributeError:
        return False


def find_product(name_prduct, fiability=0.9):
    books = make_books()
    for book in books:
        if not book[book['Producto'].apply(lambda x: ngram.NGram.compare(str(x), name_prduct, N=1) >= fiability)].empty:
            code_father = get_element_father(book)
            if fiability < 0.9:
                return [name_prduct, code_father[0], code_father[1], "Revisar", "Posible error"]
            return [name_prduct, code_father[0], code_father[1]]
    books = make_books()
    for book in books:
        if not book[book['Producto'].apply(lambda x: get_by_split(name_prduct, x))].empty:
            code_father = get_element_father(book)
            return [name_prduct, code_father[0], code_father[1], "Revisar", "Encontrado Parcialmente", ]
    if fiability > 0.5:
        return find_product(name_prduct, 0.5)


def main(column='a'):
    for factura in glob(os.path.join('FacturasPendientes', '*.xlsx')):
        doc = openpyxl.load_workbook(factura)
        for hoja in doc.sheetnames:
            enontrados = []
            noEncontrados = []
            items = tqdm(iter_row_by_column(doc[hoja], column))
            for e in items:
                items.set_description(f'Procesando {factura}-{hoja} Producto: {e}')
                result = find_product(e)
                if result:
                    enontrados.append(result)
                else:
                    noEncontrados.append(f'No se encontro: {e}')
            if enontrados:
                pd.DataFrame(enontrados).to_excel(
                    os.path.join(os.getcwd(), 'FacturasConCodigo',
                                 f'Encontrados_para_{hoja}.xlsx'))
            if noEncontrados:
                pd.DataFrame(noEncontrados).to_excel(
                    os.path.join(os.getcwd(), 'FacturasConCodigo',
                                 f'noEncontrados_para_{hoja}.xlsx'))
        shutil.move(factura, 'Facturas')


main()
