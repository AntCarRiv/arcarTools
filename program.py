import openpyxl
import os
from glob import glob

import pandas as pd
from tqdm import tqdm
from datetime import datetime

path = os.path.join(os.path.expanduser('~'), 'Documentos/DocumentosDeMuestra')
os.chdir(path)


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row if cell.value]


def iter_row_by_column(ws, c):
    datos = []
    for row in ws.iter_rows():
        for col in row:
            if col.coordinate[0].lower() == c.lower():
                if col.value:
                    datos.append(col.value)
    return datos


def make_books():
    doc = openpyxl.load_workbook('CLAVES DEL SAT 2.xlsx')
    for name in doc.sheetnames:
        a = []
        if name not in ['Filtros', 'Envios', 'Factura global']:
            for e in iter_rows(doc[name]):
                if len(e) > 2:
                    e = [e[0], str('').join(e[1:])]
                if len(e) == 2:
                    a.append(e)
                elif len(e) == 1:
                    try:
                        int(e)
                    except Exception:
                        a.append([0, e[0]])
                    else:
                        a.append(e + [""])

            c = pd.DataFrame(a, columns=['SKU', 'Producto'])

            c.to_excel(os.path.join('temps', f'{name}.xlsx'), index=False)


def search_item(n):
    docs = glob('temps/*.xlsx')
    for d in docs:
        doc = openpyxl.load_workbook(d)
        for name in doc.sheetnames:
            for e in iter_rows(doc[name]):
                try:
                    if n.lower() == e[1].lower():
                        return doc[name].cell(2, 1).value, doc[name].cell(2, 2).value
                    elif (n.encode('ascii', 'replace').decode().replace('?', '').replace(' ', '').lower() == e[
                        1].encode('ascii', 'replace').decode().replace('?', '').replace(' ', '').lower()):
                        return doc[name].cell(2, 1).value, doc[name].cell(2, 2).value
                except Exception as details:
                    pass


make_books()

doc = openpyxl.load_workbook('Libro3.xlsx')
for hoja in doc.sheetnames:
    enontrados = []
    noEncontrados = []
    print(f'Procesando para {hoja}')
    for e in tqdm(iter_row_by_column(doc[hoja], 'a')):
        result = search_item(e)
        if result:
            enontrados.append([e, result[0], result[1]])
        else:
            noEncontrados.append(f'No se encontro: {e}')
    if enontrados:
        pd.DataFrame(enontrados).to_excel(f'procesados/Encontrados_para_{hoja}_{datetime.now().isoformat()}.xlsx')
    if noEncontrados:
        pd.DataFrame(noEncontrados).to_excel(f'procesados/noEncontrados_para_{hoja}_{datetime.now().isoformat()}.xlsx')
